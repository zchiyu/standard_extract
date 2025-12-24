from __future__ import annotations

import os
import re
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


# 输出列（按你要求）
COLUMNS = (
    "order_index",   # 图编号（顺序号）
    "std_no",        # 同 toc 表
    "image_title",   # 原标题
    "clause_sort",   # 图/表
    "clause_id",     # 编号（数字与 . 或 - 交替）
    "clause_text",   # 去掉 sort+id 后剩余
    "image",         # 嵌入图片
)

# 连续 6 个及以上英文字母（大小写） -> 过滤整行
RE_LONG_ALPHA = re.compile(r"[A-Za-z]{6,}")

# clause_sort：优先开头“图/表”，否则全文找一次
RE_SORT_PREFIX = re.compile(r"^\s*([图表])")
RE_SORT_ANY = re.compile(r"([图表])")

# clause_id：紧跟在“图/表”后面的编号
# 支持：2-1、2.1、2.1.3、10-2-3
RE_ID_AFTER_SORT = re.compile(r"^[图表]\s*([0-9]+(?:[.-][0-9]+)*)")

# 兜底：任何位置出现编号也可抓（防止“图 2-1 …”中间有空格）
RE_ID_ANY = re.compile(r"([0-9]+(?:[.-][0-9]+)*)")


def _normalize_spaces(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_image_title_fields(image_title: str) -> Tuple[str, str, str]:
    """
    输入 image_title（例如：'图2-1硕士生培养流程' / '表 6 请求消息体'）
    输出 (clause_sort, clause_id, clause_text)

    clause_sort: '图' 或 '表'，否则 ''
    clause_id:   '2-1' / '6' / '2.1.3'，否则 ''
    clause_text: 去掉 sort+id 后剩余文本
    """
    t = _normalize_spaces(image_title)

    clause_sort = ""
    m = RE_SORT_PREFIX.search(t)
    if m:
        clause_sort = m.group(1)
    else:
        m2 = RE_SORT_ANY.search(t)
        if m2:
            clause_sort = m2.group(1)

    clause_id = ""
    if clause_sort:
        # 尽量用“图/表”后面的编号
        m3 = RE_ID_AFTER_SORT.search(t.replace(" ", ""))
        if m3:
            clause_id = m3.group(1)
        else:
            # 再用“图/表”后允许空格的方式
            # 例如："图 2-1 xxx"
            tmp = t
            tmp = re.sub(r"^\s*[图表]\s*", "", tmp)
            m4 = RE_ID_ANY.match(tmp)
            if m4:
                clause_id = m4.group(1)

    # clause_text：从原标题中删掉 sort 和 id，再清理残留符号
    clause_text = t

    if clause_sort:
        # 仅删除第一个出现的 sort（通常在开头）
        clause_text = re.sub(r"^\s*" + re.escape(clause_sort) + r"\s*", "", clause_text, count=1)

    if clause_id:
        # 删除开头的编号（允许前面有空格）
        clause_text = re.sub(r"^\s*" + re.escape(clause_id) + r"\s*", "", clause_text, count=1)

    # 清掉常见连接符/冒号等
    clause_text = clause_text.strip(" -—:：，,;；.。")
    clause_text = _normalize_spaces(clause_text)

    return clause_sort, clause_id, clause_text


def export_image_rows_with_embedded_images(
    rows: List[Dict[str, Any]],
    output_xlsx_path: str,
    *,
    sheet_name: str = "images",
    image_col_name: str = "image",
    image_title_col: str = "image_title",
    image_display_px: Tuple[int, int] = (320, 200),
    verbose: bool = True,
) -> str:
    """
    将 rows 导出到 image.xlsx：
      - 新增 clause_sort / clause_id / clause_text 三列（从 image_title 解析）
      - 过滤：image_title 中出现连续 >=6 英文字母的行不输出
      - 将 image 指向的图片嵌入到单元格
    """
    if not rows:
        raise ValueError("rows 为空，无法导出 image.xlsx")

    os.makedirs(os.path.dirname(output_xlsx_path) or ".", exist_ok=True)

    # 先做过滤 + 衍生字段
    filtered_rows: List[Dict[str, Any]] = []
    for r in rows:
        image_title = _normalize_spaces(str(r.get(image_title_col, "") or ""))
        if not image_title:
            # 没标题也允许输出（你也可以改成跳过）
            image_title = ""

        # 5) 过滤：>=6 连续英文字母
        if image_title and RE_LONG_ALPHA.search(image_title):
            continue

        clause_sort, clause_id, clause_text = parse_image_title_fields(image_title)

        rr = dict(r)
        rr["image_title"] = image_title
        rr["clause_sort"] = clause_sort
        rr["clause_id"] = clause_id
        rr["clause_text"] = clause_text
        filtered_rows.append(rr)

    if not filtered_rows:
        raise ValueError("过滤后 rows 为空（可能全部被英文>=6过滤），未导出任何内容")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 列宽（可调）
    col_widths = {
        "A": 12,  # order_index
        "B": 40,  # std_no
        "C": 40,  # image_title
        "D": 10,  # clause_sort
        "E": 14,  # clause_id
        "F": 50,  # clause_text
        "G": 30,  # image
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    img_w, img_h = image_display_px
    image_col_idx = COLUMNS.index(image_col_name) + 1
    image_col_letter = get_column_letter(image_col_idx)

    ok_count = 0
    miss_count = 0
    err_count = 0

    for i, r in enumerate(filtered_rows, 1):
        excel_row = i + 1

        # 写文本列（除 image）
        for col_idx, col_name in enumerate(COLUMNS, 1):
            if col_name == image_col_name:
                continue
            ws.cell(row=excel_row, column=col_idx, value=r.get(col_name, ""))

        img_path = (str(r.get(image_col_name, "") or "")).strip()
        anchor = f"{image_col_letter}{excel_row}"

        # 行高：points（≈ px * 0.75）
        ws.row_dimensions[excel_row].height = max(80, img_h * 0.75)

        if not img_path:
            ws.cell(row=excel_row, column=image_col_idx, value="")
            miss_count += 1
            continue

        if not os.path.isfile(img_path):
            ws.cell(row=excel_row, column=image_col_idx, value=f"[MISSING] {img_path}")
            miss_count += 1
            continue

        # 4) 嵌入图片
        try:
            xl_img = XLImage(img_path)
            xl_img.width = img_w
            xl_img.height = img_h
            ws.add_image(xl_img, anchor)

            # 同时写路径，便于溯源（不影响图片显示）
            ws.cell(row=excel_row, column=image_col_idx, value=img_path)
            ok_count += 1
        except Exception as e:
            ws.cell(row=excel_row, column=image_col_idx, value=f"[IMG_ERROR] {img_path} | {type(e).__name__}: {e}")
            err_count += 1

    wb.save(output_xlsx_path)

    if verbose:
        print(f"[image_excel] wrote: {output_xlsx_path}")
        print(f"[image_excel] embedded_ok={ok_count} missing={miss_count} errors={err_count}")
        print(f"[image_excel] rows_in={len(rows)} rows_out(after_filter)={len(filtered_rows)}")

    return output_xlsx_path